"""恒定市值法定投策略（简化版）：回测引擎 + Excel生成

策略要点：
- 恒定市值法：持仓 < 目标买入差额，持仓 > 目标直接卖出全部超额
- 极端清仓：偏离度超 +FULL_LIQUIDATE_THRESHOLD 全清仓；超 +PARTIAL_LIQUIDATE_THRESHOLD 减至目标50%
- 无加码机制
"""

import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from common.data import (
    TARGET_NAMES, BACKTEST_PERIODS,
    download_price_data, save_price_cache, save_backtest_cache,
)

# ═══════════════════════════════════════════════════════════
# Strategy Config
# ═══════════════════════════════════════════════════════════
BASE_AMOUNTS = [1500, 1500, 600, 2400]
WEIGHTS = [0.25, 0.25, 0.10, 0.40]
TOTAL_PER_PERIOD = 6000
PAUSE_TOTAL = 150000
INCREMENT_PER_PERIOD = 2000

PARTIAL_LIQUIDATE_THRESHOLD = 0.35   # 减至目标市值 50%，不进冷却
FULL_LIQUIDATE_THRESHOLD = 0.55      # 全仓清仓 + 冷却 + 重置期数
RESERVE_INTEREST_ANNUAL = 0.01
COOLDOWN_RESUME = 0.03
LARGE_INVEST_MULT = 2.5  # regular > base × 2.5 视为单期大额投入

STRATEGY_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Styles ──────────────────────────────────────────────
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
backtest_header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
normal_font = Font(size=11)
bold_font = Font(bold=True, size=11)
green_font = Font(size=11, color="008000")
red_font = Font(size=11, color="CC0000")
orange_font = Font(size=11, color="E67E00")
purple_font = Font(size=11, color="7030A0")
gray_font = Font(size=11, color="808080")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
large_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")          # 琥珀黄
partial_liquidate_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")  # 浅橙
liquidate_fill = PatternFill(start_color="FFDCD8", end_color="FFDCD8", fill_type="solid")      # 浅红
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
yuan_fmt = "#,##0.00"
pct_fmt = "0.00%"


def style_header_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_cell(ws, row, col, fmt=None, font=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or normal_font
    cell.alignment = center
    cell.border = thin_border
    if fmt:
        cell.number_format = fmt
    return cell


# ═══════════════════════════════════════════════════════════
# DCA Engine
# ═══════════════════════════════════════════════════════════
def compute_period(base, period, price, ma120, holding, state, global_cumulative, frozen_target=None):
    target_val = frozen_target if frozen_target is not None else base * period
    deviation = (price - ma120) / ma120 if ma120 != 0 else 0
    paused = global_cumulative >= PAUSE_TOTAL

    if state == "cooldown":
        return dict(
            target=target_val, deviation=deviation,
            regular=0, harvest=0, extra=0, actual=0,
            notes="冷却期中（偏离度{:+.1f}%，等待回落）".format(deviation * 100),
            state_out="cooldown" if deviation >= COOLDOWN_RESUME else "resume_next",
        )

    # 全仓清仓（优先级最高）
    if deviation >= FULL_LIQUIDATE_THRESHOLD and holding > 0:
        return dict(
            target=target_val, deviation=deviation,
            regular=0, harvest=0, extra=0, actual=-holding,
            notes="全仓清仓（偏离度{:+.1f}%）".format(deviation * 100),
            state_out="cooldown",
        )

    # 减半清仓：卖至目标市值的 50%，同时目标市值也减半
    partial_target = target_val * 0.5
    if deviation >= PARTIAL_LIQUIDATE_THRESHOLD and holding > partial_target:
        sell_amount = holding - partial_target
        return dict(
            target=target_val, deviation=deviation,
            regular=0, harvest=-sell_amount, extra=0, actual=-sell_amount,
            notes="减半清仓（偏离度{:+.1f}%，减至目标50%%）".format(deviation * 100),
            state_out="normal",
            partial_liquidate=True,
        )

    gap = max(target_val - holding, 0)
    excess = holding - target_val

    # Pure constant value: sell ALL excess when holding > target
    harvest = -excess if excess > 0 else 0.0
    regular = gap  # no cap: invest the full gap
    actual = regular + harvest  # harvest is negative

    notes_parts = []
    if paused:
        notes_parts.append("[增量阶段]")

    if state == "resume":
        notes_parts.append("冷却解除（偏离度{:+.1f}%），恢复定投".format(deviation * 100))
    elif period == 1 and not paused:
        notes_parts.append("初始买入")
    elif regular == 0 and harvest == 0:
        notes_parts.append("本期无操作")
    else:
        if regular > 0:
            if regular < base * 0.9:
                notes_parts.append("涨了少投")
            elif regular > base * 1.1:
                notes_parts.append("跌了多投")
            else:
                notes_parts.append("正常投入")
        if harvest < 0:
            notes_parts.append("收割超额")

    return dict(
        target=target_val, deviation=deviation,
        regular=regular, harvest=harvest, extra=0, actual=actual,
        notes="；".join(notes_parts), state_out="normal",
    )


# ═══════════════════════════════════════════════════════════
# Sheet 1: 定投参数总览
# ═══════════════════════════════════════════════════════════
def create_parameter_overview(wb):
    ws1 = wb.active
    ws1.title = "定投参数总览"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "定投参数总览（朴素恒定市值法）"
    ws1["A1"].font = Font(bold=True, size=16, color="2F5496")
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

    row = 3
    ws1.cell(row=row, column=1, value="一、标的与配比").font = Font(bold=True, size=13, color="2F5496")
    row = 4
    for i, h in enumerate(["标的名称", "市场", "权重", "每期基准金额(元)"], 1):
        ws1.cell(row=row, column=i, value=h)
    style_header_row(ws1, row, 4)
    markets = ["A股", "A股", "港股", "美股"]
    for i in range(4):
        r = row + 1 + i
        style_data_cell(ws1, r, 1).value = TARGET_NAMES[i]
        style_data_cell(ws1, r, 2).value = markets[i]
        style_data_cell(ws1, r, 3, fmt=pct_fmt).value = WEIGHTS[i]
        style_data_cell(ws1, r, 4, fmt=yuan_fmt).value = BASE_AMOUNTS[i]
    r = row + 6
    style_data_cell(ws1, r, 1, font=bold_font).value = "合计"
    style_data_cell(ws1, r, 2)
    style_data_cell(ws1, r, 3, fmt=pct_fmt, font=bold_font).value = 1.0
    style_data_cell(ws1, r, 4, fmt=yuan_fmt, font=bold_font).value = TOTAL_PER_PERIOD

    row = r + 2
    ws1.cell(row=row, column=1, value="二、恒定市值法参数").font = Font(bold=True, size=13, color="2F5496")
    for i, (k, v) in enumerate([
        ("定投频率", "每两周一次"),
        ("目标市值增长方式", "线性递增（每期 +基准金额）"),
        ("目标市值公式", "目标市值(n) = 基准金额 × n"),
        ("应投金额公式", "应投 = 目标市值(n) - 当前持仓市值"),
        ("持仓 < 目标时", "买入全部差额（无封顶）"),
        ("持仓 > 目标时", "卖出全部超额（直接收割，无档位阈值）"),
    ]):
        r = row + 1 + i
        style_data_cell(ws1, r, 1, font=bold_font).value = k
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data_cell(ws1, r, 2).value = v
        ws1.cell(row=r, column=2).alignment = left_align

    row = r + 2
    ws1.cell(row=row, column=1, value="三、清仓机制").font = Font(bold=True, size=13, color="2F5496")
    for i, (k, v) in enumerate([
        ("减半清仓条件", "偏离度超 +{:.0f}%，持仓减至目标市值 50%，不进冷却期".format(PARTIAL_LIQUIDATE_THRESHOLD * 100)),
        ("全仓清仓条件", "偏离度超 +{:.0f}%，全部卖出并进入冷却期".format(FULL_LIQUIDATE_THRESHOLD * 100)),
        ("冷却解除条件", "均线偏离度回落至 +{:.0f}% 以下，目标市值从 0 重新累积".format(COOLDOWN_RESUME * 100)),
        ("各标的独立判断", "清仓/冷却互不影响"),
    ]):
        r = row + 1 + i
        style_data_cell(ws1, r, 1, font=bold_font).value = k
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data_cell(ws1, r, 2).value = v
        ws1.cell(row=r, column=2).alignment = left_align

    row = r + 2
    ws1.cell(row=row, column=1, value="四、增量阶段机制").font = Font(bold=True, size=13, color="2F5496")
    for i, (k, v) in enumerate([
        ("触发条件", "四只标的累计总投入达到 {:,.0f} 元".format(PAUSE_TOTAL)),
        ("增量阶段行为", "常规定投和收割均从储备金池流转，保留极端清仓"),
        ("目标市值", "以全速率继续递增（每期按权重增长）"),
        ("储备金不足时", "按比例缩减各标的常规投入"),
    ]):
        r = row + 1 + i
        style_data_cell(ws1, r, 1, font=bold_font).value = k
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data_cell(ws1, r, 2).value = v
        ws1.cell(row=r, column=2).alignment = left_align

    row = r + 2
    ws1.cell(row=row, column=1, value="五、安全阀与资金管理").font = Font(bold=True, size=13, color="2F5496")
    for i, (k, v) in enumerate([
        ("单期单标的买入上限", "无封顶，按恒定市值法自然计算"),
        ("收割/清仓资金去向", "统一进入储备金池"),
        ("储备金池建议存放", "货币基金（年化 {:.0f}%，回测中按复利计息）".format(RESERVE_INTEREST_ANNUAL * 100)),
    ]):
        r = row + 1 + i
        style_data_cell(ws1, r, 1, font=bold_font).value = k
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data_cell(ws1, r, 2).value = v
        ws1.cell(row=r, column=2).alignment = left_align

    row = r + 2
    ws1.cell(row=row, column=1, value="六、增量资金方案").font = Font(bold=True, size=13, color="2F5496")
    for i, (k, v) in enumerate([
        ("增量资金", "每期 {:,.0f} 元".format(INCREMENT_PER_PERIOD)),
        ("启动条件", "存量定投达 {:,.0f} 元后自动启动".format(PAUSE_TOTAL)),
        ("资金流向", "增量资金进入储备金池，目标市值以全速率递增"),
        ("常规定投", "按目标市值差额从储备金池投出"),
        ("储备金不足", "按比例缩减各标的投入"),
    ]):
        r = row + 1 + i
        style_data_cell(ws1, r, 1, font=bold_font).value = k
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        style_data_cell(ws1, r, 2).value = v
        ws1.cell(row=r, column=2).alignment = left_align

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 42


# ═══════════════════════════════════════════════════════════
# Sheet helpers
# ═══════════════════════════════════════════════════════════
HEADERS_TARGET = [
    "日期", "期数",
    "目标市值", "当前价格", "120日均线",
    "均线偏离度", "当前持仓市值",
    "常规应投", "网格收割",
    "本期实际操作", "累计投入", "操作说明",
]
COL_WIDTHS = [12, 6, 12, 12, 12, 12, 14, 12, 12, 14, 14, 45]


def write_target_sheet(ws, tname, base, rows_data, start_row=1):
    for i, h in enumerate(HEADERS_TARGET, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(HEADERS_TARGET), fill=backtest_header_fill)

    cumulative = 0.0
    for idx, row_data in enumerate(rows_data):
        r = start_row + 1 + idx
        d = row_data

        if d["actual"] > 0:
            cumulative += d["actual"]

        style_data_cell(ws, r, 1).value = d["date"]
        style_data_cell(ws, r, 2).value = d["period"]
        style_data_cell(ws, r, 3, fmt=yuan_fmt).value = d["target"]
        style_data_cell(ws, r, 4, fmt="#,##0.0000" if d["price"] < 10 else "#,##0.00").value = d["price"]
        style_data_cell(ws, r, 5, fmt="#,##0.0000" if d["ma120"] < 10 else "#,##0.00").value = d["ma120"]
        style_data_cell(ws, r, 6, fmt="+0.00%;-0.00%").value = d["deviation"]
        style_data_cell(ws, r, 7, fmt=yuan_fmt).value = d["holding"]

        reg_font = normal_font
        if d["regular"] > base * 1.1:
            reg_font = green_font
        elif 0 < d["regular"] < base * 0.9:
            reg_font = orange_font
        style_data_cell(ws, r, 8, fmt=yuan_fmt, font=reg_font).value = d["regular"]
        style_data_cell(ws, r, 9, fmt=yuan_fmt,
                        font=red_font if d["harvest"] < 0 else normal_font).value = d["harvest"]

        if d["actual"] < 0:
            act_font = Font(bold=True, size=11, color="CC0000")
        elif d["actual"] > 0:
            act_font = bold_font
        else:
            act_font = Font(bold=True, size=11, color="808080")
        style_data_cell(ws, r, 10, fmt=yuan_fmt, font=act_font).value = d["actual"]
        style_data_cell(ws, r, 11, fmt=yuan_fmt).value = cumulative

        notes_font = normal_font
        is_liquidate = "全仓清仓" in d["notes"] and d["actual"] < 0
        is_partial_liquidate = "减半清仓" in d["notes"]
        is_large = d["regular"] > base * LARGE_INVEST_MULT

        if is_liquidate:
            notes_font = red_font
        elif is_partial_liquidate:
            notes_font = orange_font
        elif "冷却" in d["notes"]:
            notes_font = gray_font
        elif "恢复" in d["notes"] or "解除" in d["notes"]:
            notes_font = green_font
        elif "暂停" in d["notes"] or "增量" in d["notes"]:
            notes_font = purple_font

        notes_text = d["notes"]
        if is_large:
            notes_text = "【大额】" + notes_text

        style_data_cell(ws, r, 12, font=notes_font).value = notes_text
        ws.cell(row=r, column=12).alignment = left_align

        # Row highlight (applied after all cells are written so border stays)
        row_fill = (liquidate_fill if is_liquidate
                    else partial_liquidate_fill if is_partial_liquidate
                    else large_fill if is_large
                    else None)
        if row_fill:
            for col in range(1, len(HEADERS_TARGET) + 1):
                ws.cell(row=r, column=col).fill = row_fill

    return start_row + 1 + len(rows_data), cumulative


# ═══════════════════════════════════════════════════════════
# Backtest engine
# ═══════════════════════════════════════════════════════════
class TargetState:
    def __init__(self, base):
        self.base = base
        self.shares = 0.0
        self.cumulative_invested = 0.0
        self.state = "normal"
        self.period = 0
        self.harvest_count = 0
        self.extra_count = 0  # always 0, kept for plot compat
        self.liquidate_count = 0
        self.peak_value = 0.0
        self.max_drawdown = 0.0
        self.frozen_target = None
        self.period_offset = 0  # reset after each liquidation


def run_backtest(price_data, bt_start, bt_end):
    dfs_in_range = {}
    for tname in TARGET_NAMES:
        df = price_data[tname]
        subset = df.loc[bt_start:bt_end]
        if len(subset) == 0:
            return None
        dfs_in_range[tname] = subset

    common_start = max(df.index[0] for df in dfs_in_range.values())
    common_end = min(df.index[-1] for df in dfs_in_range.values())

    ref_dates = dfs_in_range[TARGET_NAMES[0]].loc[common_start:common_end].index
    sample_indices = list(range(0, len(ref_dates), 10))
    sample_dates = [ref_dates[i] for i in sample_indices]

    if len(sample_dates) < 2:
        return None

    states = {tname: TargetState(BASE_AMOUNTS[i]) for i, tname in enumerate(TARGET_NAMES)}
    global_cumulative = 0.0
    reserve_pool = 0.0
    increment_cumulative = 0.0
    increment_deployed = 0.0
    backtest_rows = {tname: [] for tname in TARGET_NAMES}

    for date in sample_dates:
        paused = global_cumulative >= PAUSE_TOTAL

        if paused:
            reserve_pool += INCREMENT_PER_PERIOD
            increment_cumulative += INCREMENT_PER_PERIOD
            reserve_pool *= 1 + RESERVE_INTEREST_ANNUAL * 14 / 365

        plans = []
        for t_idx, tname in enumerate(TARGET_NAMES):
            ts = states[tname]
            df = price_data[tname]

            if date not in df.index:
                nearest = df.index[df.index.get_indexer([date], method="nearest")[0]]
                date_use = nearest
            else:
                date_use = date

            price = float(df.loc[date_use, "close"])
            ma120 = float(df.loc[date_use, "ma120"])
            holding = ts.shares * price
            ts.period += 1
            deviation = (price - ma120) / ma120

            if paused and ts.state not in ("cooldown",):
                effective_prev = (ts.period - 1) - ts.period_offset
                if ts.frozen_target is None:
                    ts.frozen_target = ts.base * max(0, effective_prev)
                ts.frozen_target += ts.base * INCREMENT_PER_PERIOD / TOTAL_PER_PERIOD

            effective_period = ts.period - ts.period_offset

            effective_state = ts.state
            if ts.state == "cooldown" and deviation < COOLDOWN_RESUME:
                effective_state = "resume"

            result = compute_period(ts.base, effective_period, price, ma120, holding,
                                    effective_state, global_cumulative,
                                    frozen_target=ts.frozen_target)

            plans.append({
                "tname": tname, "ts": ts, "date_use": date_use,
                "price": price, "ma120": ma120, "holding": holding,
                "result": result, "deviation": deviation,
            })

        # Scale regular investments if reserve pool insufficient (paused phase)
        if paused:
            active_plans = [(i, plans[i]) for i in range(len(plans))
                            if plans[i]["result"].get("state_out") not in ("cooldown",)]
            total_regular = sum(p["result"]["regular"] for _, p in active_plans
                                if p["result"]["regular"] > 0)

            if total_regular > reserve_pool and total_regular > 0:
                scale = reserve_pool / total_regular
                for idx, p in active_plans:
                    if p["result"]["regular"] > 0:
                        p["result"]["regular"] *= scale
                        p["result"]["actual"] = (p["result"]["regular"]
                                                 + p["result"]["harvest"])
                total_regular = reserve_pool

        for p in plans:
            tname = p["tname"]
            ts = p["ts"]
            result = p["result"]
            price = p["price"]
            holding = p["holding"]
            date_use = p["date_use"]
            actual = result["actual"]

            if result.get("state_out") == "cooldown":
                if holding > 0:
                    reserve_pool += holding
                    ts.shares = 0
                    ts.liquidate_count += 1
                    ts.period_offset = ts.period  # target resets from 0 after liquidation
                    ts.frozen_target = None
                ts.state = "cooldown"
            else:
                if actual > 0:
                    shares_bought = actual / price
                    ts.shares += shares_bought
                    ts.cumulative_invested += actual
                    if paused:
                        reserve_pool -= actual
                        increment_deployed += actual
                    else:
                        global_cumulative += actual
                elif actual < 0:
                    sell_amount = abs(actual)
                    shares_sold = sell_amount / price
                    ts.shares = max(0, ts.shares - shares_sold)
                    reserve_pool += sell_amount

                # 减半清仓后，有效期数（目标市值）同步减半
                if result.get("partial_liquidate"):
                    ep = ts.period - ts.period_offset
                    ts.period_offset = ts.period - ep / 2
                    if ts.frozen_target is not None:
                        ts.frozen_target *= 0.5

                ts.state = result.get("state_out", "normal")

            current_value = ts.shares * price
            ts.peak_value = max(ts.peak_value, current_value)
            if ts.peak_value > 0:
                dd = (ts.peak_value - current_value) / ts.peak_value
                ts.max_drawdown = max(ts.max_drawdown, dd)

            notes = result["notes"]
            if paused and "[增量阶段]" not in notes:
                notes = "[增量阶段] " + notes

            backtest_rows[tname].append(dict(
                date=date_use.strftime("%Y-%m-%d"), period=ts.period,
                price=price, ma120=p["ma120"], holding=holding,
                **{k: result[k] for k in ["target", "deviation", "regular", "harvest", "extra", "actual"]},
                notes=notes,
            ))

    for tname in TARGET_NAMES:
        ts = states[tname]
        ts.harvest_count = sum(1 for r in backtest_rows[tname] if r["harvest"] < 0)

    return dict(
        states=states,
        global_cumulative=global_cumulative,
        reserve_pool=reserve_pool,
        increment_cumulative=increment_cumulative,
        increment_deployed=increment_deployed,
        backtest_rows=backtest_rows,
    )


def write_summary_sheet(wb, bt_result, price_data, label, bt_start, bt_end):
    states = bt_result["states"]
    backtest_rows = bt_result["backtest_rows"]
    global_cumulative = bt_result["global_cumulative"]
    reserve_pool = bt_result["reserve_pool"]
    increment_cumulative = bt_result.get("increment_cumulative", 0)
    increment_deployed = bt_result.get("increment_deployed", 0)

    first_date = backtest_rows[TARGET_NAMES[0]][0]["date"]
    last_date = backtest_rows[TARGET_NAMES[0]][-1]["date"]

    sheet_name = f"汇总({label})"
    ws_sum = wb.create_sheet(sheet_name)

    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"] = f"回测汇总 — {label}（{first_date} ~ {last_date}）"
    ws_sum["A1"].font = Font(bold=True, size=16, color="548235")
    ws_sum["A1"].alignment = left_align

    row = 3
    headers_sum = ["标的名称", "本金/投入(元)", "持仓市值(元)", "已回收/储备金(元)", "净资产(元)", "盈亏(元)", "收益率", "年化收益率"]
    for i, h in enumerate(headers_sum, 1):
        ws_sum.cell(row=row, column=i, value=h)
    style_header_row(ws_sum, row, len(headers_sum), fill=backtest_header_fill)

    total_invested = 0
    total_holding = 0
    n_periods = len(backtest_rows[TARGET_NAMES[0]])
    years = n_periods * 14 / 365
    for t_idx, tname in enumerate(TARGET_NAMES):
        ts = states[tname]
        r = row + 1 + t_idx

        last_row = backtest_rows[tname][-1]
        final_price = last_row["price"]
        final_value = ts.shares * final_price
        invested = ts.cumulative_invested
        harvested = sum(abs(rd["harvest"]) for rd in backtest_rows[tname] if rd["harvest"] < 0)
        liquidated = sum(abs(rd["actual"]) for rd in backtest_rows[tname]
                         if "清仓" in rd["notes"] and rd["actual"] < 0)
        total_recovered = harvested + liquidated
        total_asset = final_value + total_recovered
        pnl = total_asset - invested
        ret = pnl / invested if invested > 0 else 0

        total_invested += invested
        total_holding += final_value

        style_data_cell(ws_sum, r, 1).value = tname
        style_data_cell(ws_sum, r, 2, fmt=yuan_fmt).value = invested
        style_data_cell(ws_sum, r, 3, fmt=yuan_fmt).value = final_value
        style_data_cell(ws_sum, r, 4, fmt=yuan_fmt).value = total_recovered
        style_data_cell(ws_sum, r, 5, fmt=yuan_fmt, font=bold_font).value = total_asset
        style_data_cell(ws_sum, r, 6, fmt=yuan_fmt,
                        font=green_font if pnl >= 0 else red_font).value = pnl
        style_data_cell(ws_sum, r, 7, fmt=pct_fmt,
                        font=green_font if ret >= 0 else red_font).value = ret
        etf_annualized = (1 + ret) ** (1 / years) - 1 if years > 0 and ret > -1 else 0
        style_data_cell(ws_sum, r, 8, fmt=pct_fmt,
                        font=green_font if etf_annualized >= 0 else red_font).value = etf_annualized

    r = row + 6
    user_cash = global_cumulative + increment_cumulative
    net_worth = total_holding + reserve_pool
    total_pnl = net_worth - user_cash
    total_ret = total_pnl / user_cash if user_cash > 0 else 0
    annualized_ret = (1 + total_ret) ** (1 / years) - 1 if years > 0 else 0
    style_data_cell(ws_sum, r, 1, font=bold_font).value = "合计"
    style_data_cell(ws_sum, r, 2, fmt=yuan_fmt, font=bold_font).value = user_cash
    style_data_cell(ws_sum, r, 3, fmt=yuan_fmt, font=bold_font).value = total_holding
    style_data_cell(ws_sum, r, 4, fmt=yuan_fmt, font=bold_font).value = reserve_pool
    style_data_cell(ws_sum, r, 5, fmt=yuan_fmt, font=bold_font).value = net_worth
    pnl_font = Font(bold=True, size=11, color="008000" if total_pnl >= 0 else "CC0000")
    style_data_cell(ws_sum, r, 6, fmt=yuan_fmt, font=pnl_font).value = total_pnl
    style_data_cell(ws_sum, r, 7, fmt=pct_fmt, font=pnl_font).value = total_ret
    style_data_cell(ws_sum, r, 8, fmt=pct_fmt, font=pnl_font).value = annualized_ret

    r += 1
    ws_sum.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    note_cell = ws_sum.cell(row=r, column=1,
                            value="注：合计行以用户实际出资为本金，净资产=持仓市值+储备金余额；各标的以部署资金为基准")
    note_cell.font = Font(size=10, color="808080", italic=True)
    note_cell.alignment = left_align

    r += 2
    ws_sum.cell(row=r, column=1, value="操作统计").font = Font(bold=True, size=13, color="548235")
    r += 1
    headers_ops = ["标的名称", "总期数", "收割次数", "清仓次数"]
    for i, h in enumerate(headers_ops, 1):
        ws_sum.cell(row=r, column=i, value=h)
    style_header_row(ws_sum, r, 4, fill=backtest_header_fill)
    for t_idx, tname in enumerate(TARGET_NAMES):
        ts = states[tname]
        rr = r + 1 + t_idx
        style_data_cell(ws_sum, rr, 1).value = tname
        style_data_cell(ws_sum, rr, 2).value = ts.period
        style_data_cell(ws_sum, rr, 3).value = ts.harvest_count
        style_data_cell(ws_sum, rr, 4).value = ts.liquidate_count

    rr += 2
    style_data_cell(ws_sum, rr, 1, font=bold_font).value = "储备金池最终余额"
    style_data_cell(ws_sum, rr, 2, fmt=yuan_fmt, font=bold_font).value = reserve_pool
    rr += 1
    dca_only = global_cumulative
    style_data_cell(ws_sum, rr, 1, font=bold_font).value = "存量定投投入"
    style_data_cell(ws_sum, rr, 2, fmt=yuan_fmt, font=bold_font).value = dca_only
    rr += 1
    paused_str = "是（第 {} 期后）".format(
        next((rd["period"] for tname in TARGET_NAMES for rd in backtest_rows[tname]
              if "增量" in rd["notes"] or "暂停" in rd["notes"]), "未触发")
    ) if dca_only >= PAUSE_TOTAL else "否（总投入 {:,.0f} 元，未达 {:,.0f} 元）".format(
        dca_only, PAUSE_TOTAL)
    style_data_cell(ws_sum, rr, 1, font=bold_font).value = "是否进入增量阶段"
    ws_sum.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
    style_data_cell(ws_sum, rr, 2).value = paused_str
    ws_sum.cell(row=rr, column=2).alignment = left_align

    if increment_cumulative > 0:
        rr += 2
        ws_sum.cell(row=rr, column=1, value="增量资金统计").font = Font(bold=True, size=13, color="548235")
        rr += 1
        style_data_cell(ws_sum, rr, 1, font=bold_font).value = "增量资金（每期）"
        style_data_cell(ws_sum, rr, 2, fmt=yuan_fmt).value = INCREMENT_PER_PERIOD
        rr += 1
        style_data_cell(ws_sum, rr, 1, font=bold_font).value = "增量累计入储备金"
        style_data_cell(ws_sum, rr, 2, fmt=yuan_fmt).value = increment_cumulative
        rr += 1
        style_data_cell(ws_sum, rr, 1, font=bold_font).value = "从储备金部署投入"
        style_data_cell(ws_sum, rr, 2, fmt=yuan_fmt).value = increment_deployed

    for col_letter, width in [("A", 20), ("B", 16), ("C", 16), ("D", 16), ("E", 16),
                               ("F", 16), ("G", 14), ("H", 14)]:
        ws_sum.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Downloading historical data...")
    price_data = download_price_data()

    OUTPUT_PATH = os.path.join(STRATEGY_DIR, "定投计划.xlsx")

    print(f"\n{'='*50}")
    print("Running backtests (pure constant value strategy)")
    print(f"{'='*50}")

    wb = openpyxl.Workbook()
    create_parameter_overview(wb)

    all_results = {}

    for label, bt_start, bt_end in BACKTEST_PERIODS:
        result = run_backtest(price_data, bt_start, bt_end)
        if result is None:
            print(f"  {label}: skipped (insufficient data)")
            continue

        all_results[label] = result
        rows_all = result["backtest_rows"]
        first_date = rows_all[TARGET_NAMES[0]][0]["date"]
        last_date = rows_all[TARGET_NAMES[0]][-1]["date"]
        n_periods = len(rows_all[TARGET_NAMES[0]])
        print(f"  {label}: {first_date} ~ {last_date}, {n_periods} periods")

        for t_idx, tname in enumerate(TARGET_NAMES):
            base = BASE_AMOUNTS[t_idx]
            rows = rows_all[tname]

            short_name = tname.replace(" ETF", "")
            sheet_name = f"{short_name}({label})"
            ws = wb.create_sheet(sheet_name)

            ws.merge_cells("A1:L1")
            ws["A1"] = f"{tname}  —  回测（{rows[0]['date']} ~ {rows[-1]['date']}，共 {len(rows)} 期）"
            ws["A1"].font = Font(bold=True, size=14, color="548235")
            ws["A1"].alignment = left_align

            write_target_sheet(ws, tname, base, rows, start_row=2)

            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        write_summary_sheet(wb, result, price_data, label, bt_start, bt_end)

    wb.save(OUTPUT_PATH)
    print(f"  Saved to {OUTPUT_PATH}")

    save_price_cache(price_data)
    save_backtest_cache(all_results)
    print("  Cache saved")
