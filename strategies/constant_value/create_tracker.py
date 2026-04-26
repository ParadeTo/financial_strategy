#!/usr/bin/env python3
"""生成定投追踪 Excel 模板：参数联动、公式驱动。"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

STRATEGY_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(STRATEGY_DIR, "定投追踪记录.xlsx")

# ── 策略参数（与 strategy.py 保持一致）──────────────────
ETF_NAMES   = ["沪深300 ETF", "中证500 ETF", "恒生指数 ETF", "纳指100 ETF"]
BASE_AMOUNTS = [1500, 1500, 600, 2400]
MARKETS      = ["A股", "A股", "港股", "美股"]

PARTIAL_LIQUIDATE  = 0.35   # 减半清仓：减至目标市值 50%，不进冷却
FULL_LIQUIDATE     = 0.55   # 全仓清仓 + 冷却期
COOLDOWN_RESUME    = 0.03
PAUSE_TOTAL        = 150000   # 触发增量阶段的累计投入门槛
INCREMENT_PER_PERIOD = 2000   # 增量阶段每期注入储备金池的金额

# ── 样式 ─────────────────────────────────────────────────
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

HEADER_FILL   = _fill("2F5496")
MANUAL_FILL   = _fill("FFF2CC")   # 黄：手动填写
FORMULA_FILL  = _fill("EBF3E8")   # 绿：自动计算
SUGGEST_FILL  = _fill("DCE6F1")   # 蓝：建议参考
SUMMARY_FILL  = _fill("F4B942")   # 橙：汇总

WHITE_FONT  = Font(bold=True, size=11, color="FFFFFF")
BOLD_FONT   = Font(bold=True, size=11)
NORM_FONT   = Font(size=11)
GRAY_FONT   = Font(size=10, color="808080")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

YUAN  = '#,##0.00'
PCT   = '+0.00%;-0.00%;0.00%'
SHARE = '#,##0.0000'

def _hcell(ws, row, col, value, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = WHITE_FONT
    c.fill   = fill or HEADER_FILL
    c.alignment = CENTER
    c.border = _border()
    return c

def _cell(ws, row, col, value=None, fmt=None, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font or NORM_FONT
    c.fill      = fill or _fill("FFFFFF")
    c.alignment = align or CENTER
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c


# ═══════════════════════════════════════════════════════════
# Sheet 1: 参数配置
# ═══════════════════════════════════════════════════════════
PARAM_SHEET = "参数配置"

# Named cell positions (row in 参数配置 sheet)
PARAM_BASE_ROW        = 5   # 沪深300 base row; +etf_index for others
PARAM_PARTIAL_LIQ_ROW = 12  # 减半清仓偏离度阈值
PARAM_FULL_LIQ_ROW    = 13  # 全仓清仓偏离度阈值
# 增量阶段参数
PARAM_PAUSE_ROW  = 16  # 触发增量阶段门槛（累计净投入）
PARAM_INCR_ROW   = 17  # 增量阶段每期注入金额


def create_param_sheet(wb):
    ws = wb.active
    ws.title = PARAM_SHEET

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "定投策略参数配置（修改此表，所有记录表自动更新）"
    t.font  = Font(bold=True, size=14, color="2F5496")
    t.alignment = LEFT

    ws.merge_cells("A2:F2")
    ws["A2"].value = "黄色单元格可修改；绿色为只读说明"
    ws["A2"].font  = GRAY_FONT
    ws["A2"].alignment = LEFT

    # ── 标的配比 ────────────────────────────
    r = 4
    _hcell(ws, r, 1, "标的配比")
    for i, h in enumerate(["标的名称", "市场", "基准金额(元)", "权重"]):
        _hcell(ws, r, 2 + i, h)

    weights = [b / sum(BASE_AMOUNTS) for b in BASE_AMOUNTS]
    for i, (name, mkt, base, w) in enumerate(zip(ETF_NAMES, MARKETS, BASE_AMOUNTS, weights)):
        row = r + 1 + i
        _cell(ws, row, 1, name,   fill=FORMULA_FILL, font=BOLD_FONT)
        _cell(ws, row, 2, mkt,    fill=FORMULA_FILL)
        _cell(ws, row, 3, base,   fill=MANUAL_FILL, fmt=YUAN)
        _cell(ws, row, 4, w,      fill=FORMULA_FILL, fmt='0.0%')

    total_row = r + 5
    _cell(ws, total_row, 1, "合计", font=BOLD_FONT, fill=FORMULA_FILL)
    _cell(ws, total_row, 2, "", fill=FORMULA_FILL)
    c = _cell(ws, total_row, 3, fmt=YUAN, fill=FORMULA_FILL, font=BOLD_FONT)
    c.value = f"=SUM(C{r+1}:C{r+4})"
    c2 = _cell(ws, total_row, 4, fmt='0.0%', fill=FORMULA_FILL, font=BOLD_FONT)
    c2.value = f"=SUM(D{r+1}:D{r+4})"

    # ── 常规封顶 ────────────────────────────
    r2 = total_row + 2
    _hcell(ws, r2, 1, "清仓参数")
    _hcell(ws, r2, 2, "参数")
    _hcell(ws, r2, 3, "数值")
    _hcell(ws, r2, 4, "说明")
    rows_liq = [
        ("减半清仓偏离度阈值", PARTIAL_LIQUIDATE,
         f"偏离度超 +{int(PARTIAL_LIQUIDATE*100)}%，卖至目标市值 50%，不进冷却"),
        ("全仓清仓偏离度阈值", FULL_LIQUIDATE,
         f"偏离度超 +{int(FULL_LIQUIDATE*100)}%，全部清仓并进入冷却期"),
    ]
    for j, (k, v, desc) in enumerate(rows_liq):
        rr = r2 + 1 + j
        _cell(ws, rr, 1, k,    fill=FORMULA_FILL)
        _cell(ws, rr, 2, "",   fill=FORMULA_FILL)
        _cell(ws, rr, 3, v,    fill=MANUAL_FILL, fmt='+0%')
        _cell(ws, rr, 4, desc, fill=FORMULA_FILL, align=LEFT)

    # ── 增量阶段参数 ────────────────────────
    r3 = r2 + 4
    ws.cell(row=r3, column=1, value="三、增量阶段").font = Font(bold=True, size=13, color="2F5496")
    rows_incr = [
        ("触发门槛（累计净投入）", PAUSE_TOTAL,          "达到后切换为增量阶段，买卖均走储备金池"),
        ("每期增量注入金额",       INCREMENT_PER_PERIOD, "每期向储备金池注入的新增资金"),
    ]
    for j, (k, v, desc) in enumerate(rows_incr):
        rr = r3 + 1 + j
        _cell(ws, rr, 1, k,    fill=FORMULA_FILL, font=BOLD_FONT)
        _cell(ws, rr, 2, "",   fill=FORMULA_FILL)
        _cell(ws, rr, 3, v,    fill=MANUAL_FILL, fmt=YUAN)
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=5)
        _cell(ws, rr, 4, desc, fill=FORMULA_FILL, align=LEFT)

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════
# Sheet per ETF: 定投记录
# Column layout:
#  A  日期          手动
#  B  期数          手动
#  C  目标市值       公式  = 基准金额 × B
#  D  当前价格       手动
#  E  120日均线      手动
#  F  均线偏离度     公式  = (D-E)/E
#  G  当前持仓份额   手动（上期操作后份额，首期填0）
#  H  当前持仓市值   公式  = G × D
#  I  建议操作金额   公式  正=买 负=卖（恒定市值法：缺口补足/超额卖出/减半清仓/全清仓）
#  J  建议操作       公式  文字说明
#  K  建议金额       公式  = I（备用列，与I相同）
#  L  实际操作金额   手动  正=买入 负=卖出
#  M  实际买卖份额   手动  正=买入 负=卖出
#  N  操作后持仓份额  公式  = G + M
#  O  累计净投入     公式  = 上期O + L
#  P  储备金变化     公式  = -L（卖出补充储备金，买入减少储备金）
#  Q  备注          手动
# ═══════════════════════════════════════════════════════════

HEADERS = [
    "日期",        # A 1
    "期数",        # B 2
    "目标市值",    # C 3
    "当前价格",    # D 4
    "120日均线",   # E 5
    "均线偏离度",  # F 6
    "持仓份额",    # G 7  (上期结束后的持仓)
    "持仓市值",    # H 8
    "建议操作金额", # I 9  正=买 负=卖
    "建议操作",    # J 10
    "建议金额",    # K 11
    "实际操作(元)", # L 12  正=买  负=卖
    "买卖份额",    # M 13  正=买  负=卖
    "操作后份额",  # N 14
    "累计净投入",  # O 15
    "储备金±",    # P 16
    "备注",        # Q 17
]

COL_WIDTHS = [12, 6, 14, 12, 12, 12, 12, 14, 12, 12, 12, 14, 10, 12, 14, 12, 30]

MANUAL_COLS  = {1, 2, 4, 5, 7, 12, 13, 17}    # A B D E G L M Q
FORMULA_COLS = {3, 6, 8, 9, 14, 15, 16}        # C F H I N O P
SUGGEST_COLS = {10, 11}                         # J K


def etf_sheet_name(tname):
    return tname.replace(" ETF", "")


def create_etf_sheet(wb, tname, base_amount, etf_index, param_col):
    """
    etf_index: 0..3 对应参数表中 C..F 列（但我们在参数表里只用 C 列=基准金额，
    所以这里用 参数配置!C{PARAM_BASE_ROW + etf_index} 来引用各标的基准金额）
    param_base_cell: e.g. "参数配置!$C$5" for 沪深300
    """
    ws = wb.create_sheet(title=etf_sheet_name(tname))

    # 标题行
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    t = ws["A1"]
    t.value = f"{tname} — 定投追踪记录"
    t.font  = Font(bold=True, size=14, color="2F5496")
    t.alignment = LEFT

    # 说明行
    ws.merge_cells(f"A2:{get_column_letter(len(HEADERS))}2")
    ws["A2"].value = "黄色=手动填写  绿色=自动计算  蓝色=建议参考（可覆盖）"
    ws["A2"].font  = GRAY_FONT
    ws["A2"].alignment = LEFT

    # 图例说明
    r_legend = 3
    _cell(ws, r_legend, 1, "手动填写", fill=MANUAL_FILL,  font=GRAY_FONT, align=CENTER)
    _cell(ws, r_legend, 2, "自动计算", fill=FORMULA_FILL, font=GRAY_FONT, align=CENTER)
    _cell(ws, r_legend, 3, "建议参考", fill=SUGGEST_FILL, font=GRAY_FONT, align=CENTER)
    for col in range(4, len(HEADERS) + 1):
        _cell(ws, r_legend, col, fill=_fill("FFFFFF"))

    # 表头
    HDR_ROW = 4
    for i, h in enumerate(HEADERS, 1):
        _hcell(ws, HDR_ROW, i, h)

    # 基准金额引用（参数配置!$C$row）
    base_ref       = f"参数配置!$C${PARAM_BASE_ROW + etf_index}"
    partial_ref    = f"参数配置!$C${PARAM_PARTIAL_LIQ_ROW}"
    full_ref       = f"参数配置!$C${PARAM_FULL_LIQ_ROW}"

    # 预填 30 行数据行
    DATA_START = HDR_ROW + 1
    MAX_ROWS   = 100  # 预留100行

    for idx in range(MAX_ROWS):
        r = DATA_START + idx
        prev_r = r - 1  # for running totals

        fill_m = MANUAL_FILL
        fill_f = FORMULA_FILL
        fill_s = SUGGEST_FILL

        # A: 日期
        _cell(ws, r, 1, fill=fill_m, fmt="YYYY-MM-DD", align=CENTER)
        # B: 期数
        _cell(ws, r, 2, fill=fill_m, align=CENTER)
        # C: 目标市值 = 基准金额 × 期数
        c = _cell(ws, r, 3, fill=fill_f, fmt=YUAN)
        c.value = f"=IFERROR({base_ref}*B{r},\"\")"
        # D: 当前价格
        _cell(ws, r, 4, fill=fill_m, fmt="#,##0.0000")
        # E: 120日均线
        _cell(ws, r, 5, fill=fill_m, fmt="#,##0.0000")
        # F: 均线偏离度 = (D-E)/E
        c = _cell(ws, r, 6, fill=fill_f, fmt=PCT)
        c.value = f'=IFERROR((D{r}-E{r})/E{r},"")'
        # G: 持仓份额（手动，填上期N列的值）
        _cell(ws, r, 7, fill=fill_m, fmt=SHARE)
        # H: 持仓市值 = G × D
        c = _cell(ws, r, 8, fill=fill_f, fmt=YUAN)
        c.value = f"=IFERROR(G{r}*D{r},\"\")"
        # I: 建议操作金额（恒定市值法：正=买入差额，负=卖出超额/减半/全清）
        c = _cell(ws, r, 9, fill=fill_f, fmt=YUAN)
        c.value = (
            f'=IFERROR('
            f'IF(F{r}>={full_ref},-H{r},'
            f'IF(AND(F{r}>={partial_ref},H{r}>C{r}*0.5),-(H{r}-C{r}*0.5),'
            f'C{r}-H{r})),'
            f'"")'
        )
        # J: 建议操作
        c = _cell(ws, r, 10, fill=fill_s)
        c.value = (
            f'=IFERROR('
            f'IF(F{r}>={full_ref},"全仓清仓",'
            f'IF(AND(F{r}>={partial_ref},H{r}>C{r}*0.5),"减半清仓",'
            f'IF(H{r}>C{r},"卖出超额","正常定投"))),'
            f'"")'
        )
        # K: 建议金额 = I
        c = _cell(ws, r, 11, fill=fill_s, fmt=YUAN)
        c.value = f'=IFERROR(I{r},"")'
        # L: 实际操作金额（手动）
        _cell(ws, r, 12, fill=fill_m, fmt=YUAN)
        # M: 买卖份额（手动）
        _cell(ws, r, 13, fill=fill_m, fmt=SHARE)
        # N: 操作后份额 = G + M
        c = _cell(ws, r, 14, fill=fill_f, fmt=SHARE)
        c.value = f'=IFERROR(G{r}+M{r},"")'
        # O: 累计净投入
        c = _cell(ws, r, 15, fill=fill_f, fmt=YUAN)
        if idx == 0:
            c.value = f'=IFERROR(L{r},"")'
        else:
            c.value = f'=IFERROR(O{prev_r}+L{r},"")'
        # P: 储备金变化 = -L (买入为负，卖出为正)
        c = _cell(ws, r, 16, fill=fill_f, fmt=YUAN)
        c.value = f'=IFERROR(-L{r},"")'
        # Q: 备注
        _cell(ws, r, 17, fill=fill_m, align=LEFT)

    # 列宽
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{DATA_START}"
    ws.row_dimensions[HDR_ROW].height = 32

    return ws


# ═══════════════════════════════════════════════════════════
# Sheet: 每期汇总（增量阶段 & 储备金池跟踪）
# 列说明：
#  A  期数          手动
#  B  日期          手动
#  C  沪深300操作   公式 INDEX/MATCH 从ETF表取
#  D  中证500操作   同上
#  E  恒生指数操作  同上
#  F  纳指100操作   同上
#  G  本期自掏腰包  公式 = 存量阶段：SUM买入；增量阶段：增量注入
#  H  本期增量注入  手动 (存量阶段填0，增量阶段填2000)
#  I  本期收割入池  公式 = 各ETF卖出金额之和（负值操作的绝对值）
#  J  本期从池支出  公式 = 增量阶段买入金额（来自储备金池）
#  K  储备金池余额  公式 = 上期K + H + I - J
#  L  四标的累计净投 公式 = 上期L + 各ETF买入净额
#  M  当前阶段       公式 = IF L >= 15万 → 增量阶段，否则存量阶段
# ═══════════════════════════════════════════════════════════
def create_period_summary_sheet(wb):
    ws = wb.create_sheet(title="每期汇总")

    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = "每期汇总 — 增量阶段 & 储备金池追踪"
    t.font  = Font(bold=True, size=14, color="2F5496")
    t.alignment = LEFT

    ws.merge_cells("A2:M2")
    ws["A2"].value = (
        "黄色=手动填写  绿色=自动计算 | "
        "存量阶段：自掏腰包买入，收割进储备金；增量阶段：自掏2000/期注入池，买卖均走储备金"
    )
    ws["A2"].font  = GRAY_FONT
    ws["A2"].alignment = LEFT

    HDR_ROW = 4
    headers = [
        "期数", "日期",
        "沪深300\n操作(元)", "中证500\n操作(元)", "恒生指数\n操作(元)", "纳指100\n操作(元)",
        "本期自掏\n腰包(元)", "本期增量\n注入(元)",
        "本期收割\n入池(元)", "本期从池\n支出(元)",
        "储备金池\n余额(元)", "四标的累计\n净投入(元)", "当前阶段",
    ]
    for i, h in enumerate(headers, 1):
        _hcell(ws, HDR_ROW, i, h)

    pause_ref = f"参数配置!$C${PARAM_PAUSE_ROW}"

    DATA_START = HDR_ROW + 1
    MAX_ROWS = 100

    snames = [etf_sheet_name(n) for n in ETF_NAMES]

    for idx in range(MAX_ROWS):
        r = DATA_START + idx
        prev_r = r - 1

        # A: 期数
        _cell(ws, r, 1, fill=MANUAL_FILL, align=CENTER)
        # B: 日期
        _cell(ws, r, 2, fill=MANUAL_FILL, fmt="YYYY-MM-DD", align=CENTER)

        # C-F: 各ETF本期实际操作金额（INDEX/MATCH 按期数匹配）
        for col, sname in enumerate(snames, 3):
            c = _cell(ws, r, col, fill=FORMULA_FILL, fmt=YUAN)
            # 用 IFERROR+SUMIF 按期数汇总（同一期可能有多行，如冷却期）
            c.value = f'=IFERROR(SUMIF({sname}!$B$5:$B$1000,A{r},{sname}!$L$5:$L$1000),"")'

        # G: 本期自掏腰包
        # 存量阶段 = 所有买入之和（正值）；增量阶段 = H（增量注入额）
        c = _cell(ws, r, 7, fill=FORMULA_FILL, fmt=YUAN)
        c.value = (
            f'=IFERROR('
            f'IF(M{r}="增量阶段",H{r},'
            f'SUMPRODUCT(MAX(C{r},0),1)+SUMPRODUCT(MAX(D{r},0),1)'
            f'+SUMPRODUCT(MAX(E{r},0),1)+SUMPRODUCT(MAX(F{r},0),1)'
            f'),"")'
        )

        # H: 本期增量注入（手动）
        _cell(ws, r, 8, fill=MANUAL_FILL, fmt=YUAN)

        # I: 本期收割入池 = 各ETF卖出绝对值之和（操作为负时）
        c = _cell(ws, r, 9, fill=FORMULA_FILL, fmt=YUAN)
        c.value = (
            f'=IFERROR('
            f'MAX(-C{r},0)+MAX(-D{r},0)+MAX(-E{r},0)+MAX(-F{r},0)'
            f',"")'
        )

        # J: 本期从池支出（增量阶段的买入均来自池）
        c = _cell(ws, r, 10, fill=FORMULA_FILL, fmt=YUAN)
        c.value = (
            f'=IFERROR('
            f'IF(M{r}="增量阶段",'
            f'MAX(C{r},0)+MAX(D{r},0)+MAX(E{r},0)+MAX(F{r},0),'
            f'0),"")'
        )

        # K: 储备金池余额 = 上期K + 本期注入 + 收割入池 - 从池支出
        c = _cell(ws, r, 11, fill=FORMULA_FILL, fmt=YUAN, font=BOLD_FONT)
        if idx == 0:
            c.value = f'=IFERROR(H{r}+I{r}-J{r},"")'
        else:
            c.value = f'=IFERROR(K{prev_r}+H{r}+I{r}-J{r},"")'

        # L: 四标的累计净投入（每期净买入额累加，用于判断阶段）
        c = _cell(ws, r, 12, fill=FORMULA_FILL, fmt=YUAN)
        net = f'C{r}+D{r}+E{r}+F{r}'
        if idx == 0:
            c.value = f'=IFERROR({net},"")'
        else:
            c.value = f'=IFERROR(L{prev_r}+{net},"")'

        # M: 当前阶段
        c = _cell(ws, r, 13, fill=FORMULA_FILL, font=BOLD_FONT, align=CENTER)
        c.value = (
            f'=IFERROR('
            f'IF(L{r}>=参数配置!$C${PARAM_PAUSE_ROW},"增量阶段","存量阶段")'
            f',"")'
        )

    # 列宽
    col_ws = [6, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14, 10]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HDR_ROW].height = 40
    ws.freeze_panes = f"A{DATA_START}"

    return ws



def create_summary_sheet(wb):
    ws = wb.create_sheet(title="汇总看板", index=0)

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "定投汇总看板"
    t.font  = Font(bold=True, size=16, color="2F5496")
    t.alignment = LEFT

    ws.merge_cells("A2:H2")
    ws["A2"].value = "数据来源：各标的记录表 & 每期汇总；本表只读"
    ws["A2"].font  = GRAY_FONT
    ws["A2"].alignment = LEFT

    # ── 持仓概览 ────────────────────────────
    r = 4
    ws.cell(row=r, column=1, value="一、持仓概览").font = Font(bold=True, size=13, color="2F5496")
    r += 1
    headers = ["标的", "市场", "当前期数", "目标市值", "当前持仓市值", "累计净投入", "浮动盈亏"]
    for i, h in enumerate(headers, 1):
        _hcell(ws, r, i, h)

    total_row = r + 1 + len(ETF_NAMES)

    for idx, (tname, mkt) in enumerate(zip(ETF_NAMES, MARKETS)):
        rr = r + 1 + idx
        sname = etf_sheet_name(tname)
        last_b = f"LOOKUP(2,1/({sname}!B5:B1000<>\"\"),{sname}!B5:B1000)"
        last_c = f"LOOKUP(2,1/({sname}!C5:C1000<>\"\"),{sname}!C5:C1000)"
        last_h = f"LOOKUP(2,1/({sname}!H5:H1000<>\"\"),{sname}!H5:H1000)"
        last_o = f"LOOKUP(2,1/({sname}!O5:O1000<>\"\"),{sname}!O5:O1000)"

        _cell(ws, rr, 1, tname, fill=FORMULA_FILL, font=BOLD_FONT)
        _cell(ws, rr, 2, mkt,   fill=FORMULA_FILL)
        c = _cell(ws, rr, 3, fill=FORMULA_FILL)
        c.value = f"=IFERROR({last_b},0)"
        c = _cell(ws, rr, 4, fill=FORMULA_FILL, fmt=YUAN)
        c.value = f"=IFERROR({last_c},0)"
        c = _cell(ws, rr, 5, fill=FORMULA_FILL, fmt=YUAN)
        c.value = f"=IFERROR({last_h},0)"
        c = _cell(ws, rr, 6, fill=FORMULA_FILL, fmt=YUAN)
        c.value = f"=IFERROR({last_o},0)"
        c = _cell(ws, rr, 7, fill=FORMULA_FILL, fmt=YUAN)
        c.value = f"=IFERROR({last_h}-{last_o},0)"

    # 合计行
    for col in range(1, 8):
        fill = _fill("F8CBAD")
        if col == 1:
            _cell(ws, total_row, col, "合计", fill=fill, font=BOLD_FONT)
        elif col in [2, 3]:
            _cell(ws, total_row, col, "", fill=fill)
        else:
            c = _cell(ws, total_row, col, fill=fill, fmt=YUAN, font=BOLD_FONT)
            c.value = f"=SUM({get_column_letter(col)}{r+1}:{get_column_letter(col)}{total_row-1})"

    # ── 储备金池 & 阶段 ──────────────────────
    r2 = total_row + 2
    ws.cell(row=r2, column=1, value="二、储备金池 & 阶段").font = Font(bold=True, size=13, color="2F5496")
    r2 += 1
    kv_items = [
        ("当前阶段",
         f'=IFERROR(LOOKUP(2,1/(每期汇总!M5:M1000<>""),每期汇总!M5:M1000),"存量阶段")',
         None),
        ("储备金池余额",
         f'=IFERROR(LOOKUP(2,1/(每期汇总!K5:K1000<>""),每期汇总!K5:K1000),0)',
         YUAN),
        ("四标的累计净投入",
         f'=IFERROR(LOOKUP(2,1/(每期汇总!L5:L1000<>""),每期汇总!L5:L1000),0)',
         YUAN),
        ("距触发增量阶段还差",
         f'=MAX(0,参数配置!$C${PARAM_PAUSE_ROW}-IFERROR(LOOKUP(2,1/(每期汇总!L5:L1000<>""),每期汇总!L5:L1000),0))',
         YUAN),
    ]
    for j, (label, formula, fmt) in enumerate(kv_items):
        rr = r2 + j
        _cell(ws, rr, 1, label, fill=FORMULA_FILL, font=BOLD_FONT)
        c = _cell(ws, rr, 2, fill=FORMULA_FILL, fmt=fmt or "@", font=BOLD_FONT)
        c.value = formula
        for col in range(3, 8):
            _cell(ws, rr, col, fill=_fill("FFFFFF"))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    ws.freeze_panes = "A3"
    return ws


# ═══════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    create_param_sheet(wb)

    for i, (tname, base) in enumerate(zip(ETF_NAMES, BASE_AMOUNTS)):
        create_etf_sheet(wb, tname, base, i, param_col=3)

    create_period_summary_sheet(wb)
    create_summary_sheet(wb)

    # sheet 顺序：汇总看板 → 每期汇总 → 参数配置 → 各ETF
    wb.move_sheet("汇总看板", offset=-wb.sheetnames.index("汇总看板"))
    wb.move_sheet("每期汇总", offset=1 - wb.sheetnames.index("每期汇总"))
    wb.move_sheet(PARAM_SHEET, offset=2 - wb.sheetnames.index(PARAM_SHEET))

    wb.save(OUT_PATH)
    print(f"已生成：{OUT_PATH}")


if __name__ == "__main__":
    main()
