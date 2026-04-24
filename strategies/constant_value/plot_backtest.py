#!/usr/bin/env python3
"""Generate backtest summary charts from cached data."""

import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from strategies.constant_value.strategy import TargetState
from common.data import load_price_cache, load_backtest_cache
from common.plot import (
    plt, font_prop, ETF_NAMES, ETF_COLORS, PERIOD_COLORS, short_label, wan_func_formatter,
)
import numpy as np
import openpyxl

price_data = load_price_cache()
all_results = load_backtest_cache()

STRATEGY_DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(STRATEGY_DIR, 'charts')
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(os.path.join(STRATEGY_DIR, '定投计划.xlsx'), data_only=True)

ALL_PERIODS = list(all_results.keys())
GROUP_2Y = [p for p in ALL_PERIODS if p.startswith('2年')]
GROUP_5Y = [p for p in ALL_PERIODS if p.startswith('5年')]
GROUP_10Y = [p for p in ALL_PERIODS if p.startswith('10年')]
GROUP_20Y = [p for p in ALL_PERIODS if p.startswith('20年')]

GROUPS = [('2年窗口', GROUP_2Y), ('5年窗口', GROUP_5Y), ('10年窗口', GROUP_10Y), ('20年窗口', GROUP_20Y)]
GROUPS = [(n, g) for n, g in GROUPS if g]

data = {}
for period in ALL_PERIODS:
    sheet_name = f'汇总({period})'
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    period_data = {}
    found_stats = False
    for row in rows:
        if row[0] == '操作统计':
            found_stats = True
            continue
        if not found_stats:
            if row[0] in ETF_NAMES:
                period_data[row[0]] = {
                    'invested': row[1] or 0,
                    'holding': row[2] or 0,
                    'recovered': row[3] or 0,
                    'total_asset': row[4] or 0,
                    'return_rate': (row[6] or 0) * 100,
                    'annualized_rate': (row[7] or 0) * 100,
                }
            elif row[0] == '合计':
                period_data['合计'] = {
                    'invested': row[1] or 0,
                    'holding': row[2] or 0,
                    'recovered': row[3] or 0,
                    'total_asset': row[4] or 0,
                    'return_rate': (row[6] or 0) * 100,
                    'annualized_rate': (row[7] or 0) * 100,
                }
        else:
            if row[0] in ETF_NAMES:
                period_data[row[0]]['harvest_count'] = row[2] or 0
                period_data[row[0]]['extra_count'] = row[3] or 0
                period_data[row[0]]['liquidate_count'] = row[4] or 0
    data[period] = period_data

# ── Chart 1: Overall returns ──
fig, ax = plt.subplots(figsize=(18, 7))
x_labels = []
rates = []
ann_rates = []
group_boundaries = []
x_colors = []

pos = 0
for group_name, group in GROUPS:
    if pos > 0:
        pos += 0.5
    group_boundaries.append((pos, pos + len(group) - 1, group_name))
    for p in group:
        x_labels.append(short_label(p))
        rates.append(data[p]['合计']['return_rate'])
        ann_rates.append(data[p]['合计']['annualized_rate'])
        key = p[:2]
        x_colors.append(PERIOD_COLORS.get(key, '#457B9D'))
        pos += 1

x_pos = []
pos = 0
for i, (_, group) in enumerate(GROUPS):
    if i > 0:
        pos += 0.5
    for _ in group:
        x_pos.append(pos)
        pos += 1

bars = ax.bar(x_pos, rates, color=x_colors, width=0.7, edgecolor='white', linewidth=1.5)
for bar, rate, ann in zip(bars, rates, ann_rates):
    sign = '+' if rate >= 0 else ''
    y = bar.get_height()
    if y < 0:
        y_text = y - 2
        va = 'top'
    else:
        y_text = y + 0.5
        va = 'bottom'
    ax.text(bar.get_x() + bar.get_width()/2, y_text,
            f'{sign}{rate:.1f}%', ha='center', va=va, fontsize=13, fontweight='bold',
            fontproperties=font_prop)
    mid = bar.get_height() * 0.5 if bar.get_height() > 15 else bar.get_height() + 5
    ax.text(bar.get_x() + bar.get_width()/2, mid,
            f'年化{ann:.1f}%', ha='center', va='center', fontsize=10,
            color='white' if bar.get_height() > 15 else '#333',
            fontweight='bold', fontproperties=font_prop)

for start, end, name in group_boundaries:
    mid = (start + end) / 2
    ax.text(mid, -max(abs(r) for r in rates) * 0.08, name, ha='center', va='top',
            fontsize=14, fontweight='bold', fontproperties=font_prop, color='#555')

ax.set_xticks(x_pos)
ax.set_xticklabels(x_labels, fontsize=12, fontproperties=font_prop)
ax.set_ylabel('收益率 (%)', fontsize=14, fontproperties=font_prop)
ax.set_title('多窗口回测：整体收益率对比（以本金为基准）', fontsize=20, fontweight='bold',
             fontproperties=font_prop, pad=15)
min_rate = min(rates)
max_rate = max(rates)
ax.set_ylim(min(0, min_rate * 1.3), max_rate * 1.3)
ax.axhline(y=0, color='gray', linewidth=0.5)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.tick_params(axis='both', labelsize=12)
for label in ax.get_yticklabels():
    label.set_fontproperties(font_prop)
fig.tight_layout()
fig.savefig(f'{OUT}/1_overall_returns.png', dpi=150, bbox_inches='tight')
plt.close()
print('Chart 1 saved')

# ── Chart 2: Per-ETF returns ──
n_groups = len(GROUPS)
fig, axes = plt.subplots(1, n_groups, figsize=(6 * n_groups, 7))
if n_groups == 1:
    axes = [axes]
for ax_idx, (group_name, group) in enumerate(GROUPS):
    ax = axes[ax_idx]
    x = np.arange(len(group))
    width = 0.18
    for i, etf in enumerate(ETF_NAMES):
        etf_rates = [data[p][etf]['return_rate'] for p in group]
        offset = (i - 1.5) * width
        bars = ax.bar(x + offset, etf_rates, width, label=etf, color=ETF_COLORS[etf],
                      edgecolor='white', linewidth=0.8)
        for bar, rate in zip(bars, etf_rates):
            sign = '+' if rate >= 0 else ''
            y = bar.get_height()
            va = 'bottom' if y >= 0 else 'top'
            ax.text(bar.get_x() + bar.get_width()/2, y + (0.5 if y >= 0 else -0.5),
                    f'{sign}{rate:.0f}%', ha='center', va=va, fontsize=8, fontproperties=font_prop)
    ax.set_xticks(x)
    ax.set_xticklabels([short_label(p) for p in group], fontsize=12, fontproperties=font_prop)
    ax.set_title(group_name, fontsize=16, fontweight='bold', fontproperties=font_prop)
    ax.axhline(y=0, color='gray', linewidth=0.5)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    for label in ax.get_yticklabels():
        label.set_fontproperties(font_prop)
    if ax_idx == 0:
        ax.legend(prop=font_prop, fontsize=10, loc='best')
        ax.set_ylabel('收益率 (%)', fontsize=14, fontproperties=font_prop)
fig.suptitle('各标的在不同回测窗口的收益率', fontsize=20, fontweight='bold',
             fontproperties=font_prop, y=1.02)
fig.tight_layout()
fig.savefig(f'{OUT}/2_per_etf_returns.png', dpi=150, bbox_inches='tight')
plt.close()
print('Chart 2 saved')

# ── Chart 3: Capital breakdown ──
n_periods_total = len(ALL_PERIODS)
cols = min(5, n_periods_total)
rows_grid = (n_periods_total + cols - 1) // cols
fig, axes = plt.subplots(rows_grid, cols, figsize=(5 * cols, 5.5 * rows_grid), sharey=False)
if rows_grid == 1:
    axes = [axes]
axes_flat = [axes[r][c] if rows_grid > 1 else axes[c]
             for r in range(rows_grid) for c in range(cols)]

for idx, period in enumerate(ALL_PERIODS):
    ax = axes_flat[idx]
    d = data[period]['合计']
    categories = ['本金', '持仓', '储备金', '净资产']
    values = [d['invested'], d['holding'], d['recovered'], d['total_asset']]
    colors_bar = ['#6c757d', '#E63946', '#2A9D8F', '#264653']
    bars = ax.bar(categories, values, color=colors_bar, width=0.6, edgecolor='white', linewidth=1)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(values) * 0.02,
                f'{val/10000:.1f}万', ha='center', va='bottom', fontsize=9, fontproperties=font_prop)
    ax.set_title(period, fontsize=13, fontweight='bold', fontproperties=font_prop)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.set_ylim(0, max(values) * 1.2)
    ax.tick_params(axis='x', labelsize=9)
    ax.tick_params(axis='y', labelsize=9)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontproperties(font_prop)
    plt.setp(ax.get_xticklabels(), rotation=25, ha='right')

for idx in range(len(ALL_PERIODS), len(axes_flat)):
    axes_flat[idx].set_visible(False)

fig.suptitle('资金构成对比（本金 vs 持仓 vs 储备金 vs 净资产）', fontsize=18, fontweight='bold',
             fontproperties=font_prop, y=1.02)
fig.tight_layout()
fig.savefig(f'{OUT}/3_capital_breakdown.png', dpi=150, bbox_inches='tight')
plt.close()
print('Chart 3 saved')

# ── Chart 4: Annualized returns ──
fig, ax = plt.subplots(figsize=(20, 7))
pos = 0
x_pos = []
for i, (_, group) in enumerate(GROUPS):
    if i > 0:
        pos += 0.5
    for _ in group:
        x_pos.append(pos)
        pos += 1

ann_rates_all = [data[p]['合计']['annualized_rate'] for p in ALL_PERIODS]
bars = ax.bar(x_pos, ann_rates_all, color=[PERIOD_COLORS.get(p[:2], '#457B9D') for p in ALL_PERIODS],
              width=0.7, edgecolor='white', linewidth=1.5)
for bar, rate, p in zip(bars, ann_rates_all, ALL_PERIODS):
    sign = '+' if rate >= 0 else ''
    y = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2, y + 0.3,
            f'{sign}{rate:.1f}%', ha='center', va='bottom', fontsize=14, fontweight='bold',
            fontproperties=font_prop)

cum_idx = 0
for group_name, group in GROUPS:
    if not group:
        continue
    avg = np.mean([data[p]['合计']['annualized_rate'] for p in group])
    start_idx = cum_idx
    end_idx = cum_idx + len(group)
    xmin = x_pos[start_idx] - 0.4
    xmax = x_pos[end_idx - 1] + 0.4
    ax.hlines(avg, xmin, xmax, colors='red', linewidth=1.5, linestyles='--', alpha=0.7)
    ax.text(xmax + 0.1, avg, f'均值{avg:.1f}%', fontsize=11, color='red',
            fontproperties=font_prop, va='center')
    cum_idx = end_idx

ax.set_xticks(x_pos)
ax.set_xticklabels([short_label(p) for p in ALL_PERIODS], fontsize=12, fontproperties=font_prop)
ax.set_ylabel('年化收益率 (%)', fontsize=14, fontproperties=font_prop)
ax.set_title('多窗口回测：年化收益率对比', fontsize=20, fontweight='bold',
             fontproperties=font_prop, pad=15)
ax.axhline(y=0, color='gray', linewidth=0.5)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
for label in ax.get_yticklabels():
    label.set_fontproperties(font_prop)
fig.tight_layout()
fig.savefig(f'{OUT}/4_annualized_returns.png', dpi=150, bbox_inches='tight')
plt.close()
print('Chart 4 saved')

# ── Chart 5: Portfolio timelines ──
import strategies.constant_value.strategy as strat

timeline_periods = [p for p in ALL_PERIODS]  # include all windows

for period_key in timeline_periods:
    bt = all_results[period_key]
    backtest_rows = bt['backtest_rows']
    etf_names_list = list(ETF_NAMES)
    n_periods = len(backtest_rows[etf_names_list[0]])

    dates = []
    user_cash_list = []
    holding_list = []
    reserve_list = []
    net_worth_list = []

    global_cum = 0.0
    reserve_pool = 0.0
    increment_cum = 0.0

    for i in range(n_periods):
        date = backtest_rows[etf_names_list[0]][i]['date']
        dates.append(date)

        paused = global_cum >= strat.PAUSE_TOTAL
        if paused:
            reserve_pool += strat.INCREMENT_PER_PERIOD
            increment_cum += strat.INCREMENT_PER_PERIOD
            reserve_pool *= 1 + strat.RESERVE_INTEREST_ANNUAL * 14 / 365

        total_holding = 0
        for tname in etf_names_list:
            rows_to_i = backtest_rows[tname][:i+1]
            row = backtest_rows[tname][i]
            price = row['price']
            shares = 0
            for r in rows_to_i:
                if '清仓' in r['notes'] and r['actual'] < 0:
                    shares = 0
                elif r['actual'] > 0:
                    shares += r['actual'] / r['price']
                elif r['actual'] < 0:
                    shares -= abs(r['actual']) / r['price']
                    shares = max(0, shares)
            total_holding += shares * price

            r_data = backtest_rows[tname][i]
            actual = r_data['actual']
            if actual > 0:
                if paused:
                    reserve_pool -= actual
                else:
                    global_cum += actual
            elif actual < 0:
                if '清仓' in r_data['notes']:
                    reserve_pool += abs(actual)
                else:
                    reserve_pool += abs(r_data['harvest']) if r_data['harvest'] < 0 else 0

        user_cash = global_cum + increment_cum
        user_cash_list.append(user_cash)
        holding_list.append(total_holding)
        reserve_list.append(max(reserve_pool, 0))
        net_worth_list.append(total_holding + max(reserve_pool, 0))

    # ── Statistics ───────────────────────────────────────────────────────────
    pnl_list = [nw - uc for nw, uc in zip(net_worth_list, user_cash_list)]
    pnl_rate_list = [pnl / uc * 100 if uc > 0 else 0.0
                     for pnl, uc in zip(pnl_list, user_cash_list)]

    # Max drawdown of net_worth curve (peak-to-trough / peak)
    peak_nw = 0.0; peak_idx = 0
    max_dd = 0.0; max_dd_start = 0; max_dd_end = 0
    for i, nw in enumerate(net_worth_list):
        if nw > peak_nw:
            peak_nw = nw; peak_idx = i
        if peak_nw > 0:
            dd = (peak_nw - nw) / peak_nw
            if dd > max_dd:
                max_dd = dd; max_dd_start = peak_idx; max_dd_end = i

    max_loss_val = min(pnl_list)
    max_loss_idx = pnl_list.index(max_loss_val)
    max_loss_rate = pnl_rate_list[max_loss_idx]

    max_profit_val = max(pnl_list)
    max_profit_idx = pnl_list.index(max_profit_val)

    max_ret = max(pnl_rate_list)
    max_ret_idx = pnl_rate_list.index(max_ret)

    # ── Figure layout (2 rows: asset curves + pnl-rate) ──────────────────────
    is_20y = period_key.startswith('20年')
    is_10y = period_key.startswith('10年')
    is_2y  = period_key.startswith('2年')
    width  = 22 if is_20y else 18 if is_10y else 14
    height = 13 if not is_2y else 11

    fig, (ax, ax2) = plt.subplots(
        2, 1, figsize=(width, height),
        gridspec_kw={'height_ratios': [3, 1.5], 'hspace': 0.45},
    )

    # ── Main: asset curves ────────────────────────────────────────────────────
    ax.fill_between(range(len(dates)), net_worth_list, alpha=0.12, color='#264653')
    ax.plot(range(len(dates)), net_worth_list, color='#264653', linewidth=2.5, label='净资产（持仓+储备金）')
    ax.plot(range(len(dates)), user_cash_list, color='#6c757d', linewidth=2,
            label='本金（用户出资）', linestyle='--')
    ax.plot(range(len(dates)), holding_list, color='#E63946', linewidth=1.5, label='持仓市值', alpha=0.8)
    ax.plot(range(len(dates)), reserve_list, color='#2A9D8F', linewidth=1.5, label='储备金余额', alpha=0.8)

    # Mark max profit point
    ax.scatter([max_profit_idx], [net_worth_list[max_profit_idx]],
               color='#2A9D8F', s=120, zorder=6, marker='^')
    profit_y = net_worth_list[max_profit_idx]
    ax.annotate(f'最高收益额\n+{max_profit_val/10000:.1f}万 ({dates[max_profit_idx][:7]})',
                xy=(max_profit_idx, profit_y),
                xytext=(max_profit_idx - len(dates) // 6, profit_y * 0.88),
                fontsize=10, color='#2A9D8F', fontproperties=font_prop,
                arrowprops=dict(arrowstyle='->', color='#2A9D8F', lw=1.5),
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.85,
                          edgecolor='#2A9D8F', lw=1))

    # Mark max drawdown: shade region + text at center of drawdown
    if max_dd > 0.005:
        ax.axvspan(max_dd_start, max_dd_end, alpha=0.10, color='#c0392b')
        dd_peak_y = net_worth_list[max_dd_start]
        dd_trough_y = net_worth_list[max_dd_end]
        dd_center_x = (max_dd_start + max_dd_end) / 2
        ax.text(dd_center_x, (dd_peak_y + dd_trough_y) / 2,
                f'最大回撤\n-{max_dd:.1%}',
                ha='center', va='center', fontsize=10, color='#c0392b',
                fontproperties=font_prop,
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.88,
                          edgecolor='#c0392b', lw=1))

    # Mark max loss on main chart (if underwater)
    if max_loss_val < 0:
        loss_y = net_worth_list[max_loss_idx]
        ax.scatter([max_loss_idx], [loss_y], color='#E63946', s=100, zorder=6, marker='v')
        ax.annotate(f'最大亏损额\n{max_loss_val/10000:.1f}万 ({dates[max_loss_idx][:7]})',
                    xy=(max_loss_idx, loss_y),
                    xytext=(max_loss_idx + len(dates) // 8, loss_y * 0.92),
                    fontsize=10, color='#E63946', fontproperties=font_prop,
                    arrowprops=dict(arrowstyle='->', color='#E63946', lw=1.5),
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.85,
                              edgecolor='#E63946', lw=1))

    tick_step = max(1, len(dates) // 12)
    ax.set_xticks(range(0, len(dates), tick_step))
    ax.set_xticklabels([dates[i] for i in range(0, len(dates), tick_step)],
                        rotation=30, ha='right', fontsize=11, fontproperties=font_prop)
    ax.set_ylabel('金额 (元)', fontsize=14, fontproperties=font_prop)
    ax.set_title(f'{period_key} 回测：资金走势图', fontsize=20, fontweight='bold',
                 fontproperties=font_prop, pad=15)
    ax.legend(prop=font_prop, fontsize=12, loc='upper left')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    for label in ax.get_yticklabels():
        label.set_fontproperties(font_prop)
    ax.yaxis.set_major_formatter(wan_func_formatter)

    # Stats box — below legend (upper left)
    ml_str = f'{max_loss_val/10000:.1f}万' if max_loss_val < 0 else '无亏损'
    stats_text = (f'最大回撤率：{-max_dd:.1%}\n'
                  f'最大亏损额：{ml_str}\n'
                  f'最高收益率：+{max_ret:.1f}%\n'
                  f'最高收益额：+{max_profit_val/10000:.1f}万')
    fig.canvas.draw()
    legend_bbox = ax.get_legend().get_window_extent(fig.canvas.get_renderer())
    legend_bbox_ax = legend_bbox.transformed(ax.transAxes.inverted())
    ax.text(0.01, legend_bbox_ax.y0 - 0.02, stats_text, transform=ax.transAxes,
            fontsize=11, fontproperties=font_prop,
            verticalalignment='top', horizontalalignment='left',
            bbox=dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.90,
                      edgecolor='#888', linewidth=1.2))

    # ── PnL-rate subplot ──────────────────────────────────────────────────────
    ax2.axhline(y=0, color='gray', linewidth=1)
    ax2.fill_between(range(len(dates)), pnl_rate_list, 0,
                     where=[v >= 0 for v in pnl_rate_list], alpha=0.25, color='#2A9D8F')
    ax2.fill_between(range(len(dates)), pnl_rate_list, 0,
                     where=[v < 0 for v in pnl_rate_list], alpha=0.25, color='#E63946')
    ax2.plot(range(len(dates)), pnl_rate_list, color='#264653', linewidth=2)

    # Mark max return rate
    ax2.scatter([max_ret_idx], [max_ret], color='#2A9D8F', s=90, zorder=6, marker='^')
    ret_offset_x = -len(dates) // 7 if max_ret_idx > len(dates) // 2 else len(dates) // 7
    ax2.annotate(f'+{max_ret:.1f}% ({dates[max_ret_idx][:7]})',
                 xy=(max_ret_idx, max_ret),
                 xytext=(max(1, max_ret_idx + ret_offset_x), max_ret * 0.75),
                 fontsize=10, color='#2A9D8F', fontproperties=font_prop,
                 arrowprops=dict(arrowstyle='->', color='#2A9D8F', lw=1.5),
                 bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.85,
                           edgecolor='#2A9D8F', lw=1))

    # Mark max loss rate (only if clearly negative)
    if max_loss_rate < -1.0:
        ax2.scatter([max_loss_idx], [max_loss_rate], color='#E63946', s=90, zorder=6, marker='v')
        loss_offset_x = len(dates) // 7 if max_loss_idx < len(dates) // 2 else -len(dates) // 7
        ax2.annotate(f'{max_loss_rate:.1f}% ({dates[max_loss_idx][:7]})',
                     xy=(max_loss_idx, max_loss_rate),
                     xytext=(max(1, min(len(dates) - 2, max_loss_idx + loss_offset_x)),
                             max_loss_rate * 0.65),
                     fontsize=10, color='#E63946', fontproperties=font_prop,
                     arrowprops=dict(arrowstyle='->', color='#E63946', lw=1.5),
                     bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.85,
                               edgecolor='#E63946', lw=1))

    ax2.set_xticks(range(0, len(dates), tick_step))
    ax2.set_xticklabels([dates[i] for i in range(0, len(dates), tick_step)],
                         rotation=30, ha='right', fontsize=11, fontproperties=font_prop)
    ax2.set_ylabel('盈亏率 (%)', fontsize=13, fontproperties=font_prop)
    ax2.set_title('净资产相对本金的盈亏率走势', fontsize=14, fontweight='bold',
                  fontproperties=font_prop)
    ax2.spines['top'].set_visible(False)
    ax2.spines['right'].set_visible(False)
    for label in ax2.get_yticklabels():
        label.set_fontproperties(font_prop)

    fig.tight_layout()
    slug = period_key.replace('(', '_').replace(')', '')
    fig.savefig(f'{OUT}/5_timeline_{slug}.png', dpi=150, bbox_inches='tight')
    plt.close()
    print(f'Saved: 5_timeline_{slug}.png')

print(f'\nAll charts saved to {OUT}/')
